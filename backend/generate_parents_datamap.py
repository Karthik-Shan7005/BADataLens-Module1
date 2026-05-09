"""
Auto-generate the DataLens datamap for "Parents Study.sav".

Reads SPSS metadata via pyreadstat and applies naming-convention heuristics
to classify variables into: single, multi, scale — then emits the
question_id / variable_name / question_text / question_type / answer_option
format expected by datamap_parser.py.
"""

import re
import pyreadstat
import openpyxl
from openpyxl.styles import Font, PatternFill

SPSS_PATH = r"C:\Users\KarthikShanmugam\ClaudePOC\DataLens\Module 1\data\spss\Parents Study.sav"
OUT_PATH  = r"C:\Users\KarthikShanmugam\ClaudePOC\DataLens\Module 1\data\spss\Parents Study - Datamap (Generated).xlsx"

# ── variable inclusion: only survey question variables ─────────────────────
# Only include variables whose names start with a screener (S0, S1) or a
# question code (Q, AQ). Everything else is a system/tracking variable.
INCLUDE_PREFIXES = ("S0", "S1", "Q", "AQ")

SKIP_SUFFIXES = ("oe",)           # open-ended verbatim text
SKIP_GRID_RE  = re.compile(r"r\d+c\d+")   # grid cells like Q1_5_3r1c1
SKIP_EXACT_NAMES = {"Qlang"}   # language selector, not a survey question

SKIP_PREFIXES = (
    "Q1_1r",          # numeric count-of-children open entries (not categorical)
    "hQ",             # hidden variables
    "noanswerQ",      # no-answer flag variables
)


def should_skip(var: str) -> bool:
    # Must start with a survey question prefix
    if not var.startswith(INCLUDE_PREFIXES):
        return True
    if var in SKIP_EXACT_NAMES:
        return True
    if var.startswith(SKIP_PREFIXES):
        return True
    if var.endswith(SKIP_SUFFIXES):
        return True
    if SKIP_GRID_RE.search(var):
        return True
    return False


# ── helpers ────────────────────────────────────────────────────────────────
MULTI_SUFFIX_RE = re.compile(r"^(.+?)(r\d+)$")

def base_and_opt(var: str):
    """Return (base_question_id, True_if_multi_suffix) or (var, False)."""
    m = MULTI_SUFFIX_RE.match(var)
    if m:
        return m.group(1), True
    return var, False


def clean_label(raw: str) -> str:
    """Strip the 'VarName: ' prefix SPSS sometimes prepends."""
    if ": " in raw:
        return raw.split(": ", 1)[1].strip()
    return raw.strip()


def option_label_from_var_label(var_label: str, base: str) -> str:
    """
    For multi vars the SPSS label is usually:
      'Option text - Question text'
    or just 'Option text'
    Extract the option part.
    """
    cleaned = clean_label(var_label)
    if " - " in cleaned:
        return cleaned.split(" - ")[0].strip()
    return cleaned


def question_text_from_var_label(var_label: str) -> str:
    """Extract the question text (part after ' - ')."""
    cleaned = clean_label(var_label)
    if " - " in cleaned:
        return cleaned.split(" - ", 1)[1].strip()
    return cleaned


# ── main ──────────────────────────────────────────────────────────────────
_, meta = pyreadstat.read_sav(SPSS_PATH)

col_labels   = dict(zip(meta.column_names, meta.column_labels))
value_labels = meta.variable_value_labels   # {var: {code_float: label}}

# First pass: bucket variables into single-only OR multi groups
multi_groups: dict[str, list[str]] = {}   # base → [var, ...]
single_vars:  list[str] = []

for var in meta.column_names:
    if should_skip(var):
        continue
    base, is_multi_suffix = base_and_opt(var)
    if is_multi_suffix:
        multi_groups.setdefault(base, []).append(var)
    else:
        single_vars.append(var)

# Variables that appear as the ONLY member of their multi-group are actually
# standalone single questions (e.g. S0_3r2oe was already skipped, leaving S0_3 alone)
# — handle them as singles.
promote_to_single = []
for base, members in list(multi_groups.items()):
    if len(members) == 1:
        promote_to_single.append(base)
        single_vars.append(members[0])
for base in promote_to_single:
    del multi_groups[base]

# ── build rows ─────────────────────────────────────────────────────────────
rows = []

# Single questions — one row per question.
# SPSS enrichment in projects.py handles code→label mapping at upload time,
# so we don't need to repeat answer options here.
for var in single_vars:
    raw_label = col_labels.get(var, var)
    q_text = clean_label(raw_label)
    rows.append([var, var, q_text, "single", "", ""])

# Multi-select questions
for base in sorted(multi_groups.keys()):
    members = sorted(multi_groups[base])   # keep r1, r2, r3 ... order
    # Derive question text from the first member's label (part after " - ")
    first_label = col_labels.get(members[0], members[0])
    q_text = question_text_from_var_label(first_label)

    for var in members:
        raw_label = col_labels.get(var, var)
        opt_label = option_label_from_var_label(raw_label, base)
        rows.append([base, var, q_text, "multi", opt_label, ""])

# ── write Excel ───────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Datamap"

header = ["question_id", "variable_name", "question_text", "question_type", "answer_option", "is_weight"]
ws.append(header)

# Style header
for cell in ws[1]:
    cell.font  = Font(bold=True, color="FFFFFF")
    cell.fill  = PatternFill("solid", fgColor="4472C4")

for row in rows:
    ws.append(row)

# Auto column widths
for col in ws.columns:
    max_len = max(len(str(cell.value or "")) for cell in col)
    ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 80)

wb.save(OUT_PATH)
print(f"Saved: {OUT_PATH}")
print(f"Total rows: {len(rows)}")
print(f"  Single/scale questions: {len(single_vars)}")
print(f"  Multi-select groups   : {len(multi_groups)}")
