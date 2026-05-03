import openpyxl
import sys

output_file = r"C:\Users\KarthikShanmugam\ClaudePOC\DataLens\Module 1\data\spss\spss_structure.txt"
sys.stdout = open(output_file, "w", encoding="utf-8")

file_path = r"C:\Users\KarthikShanmugam\ClaudePOC\DataLens\Module 1\data\spss\Sample Data - DataLens.xlsx"

wb = openpyxl.load_workbook(file_path)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== Sheet: {sheet_name} ===")
    print(f"Dimensions: {ws.dimensions}")
    print("\nFirst 5 rows (raw):")
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= 5:
            break
        print(f"  Row {i+1}: {row}")

    print("\nAll rows (filling merged cells):")
    # Read with openpyxl handling merged cells
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(row)

    # Get headers from row 1
    headers = data[0]
    print(f"  Headers: {headers}")

    # Forward-fill None values in first column (merged cells)
    prev = [None] * len(headers)
    for row in data[1:]:
        filled = []
        for i, val in enumerate(row):
            if val is None:
                filled.append(prev[i])
            else:
                filled.append(val)
                prev[i] = val

        # Skip "NO TO:" rows (value 0 or label starts with "NO TO")
        row_str = str(filled)
        if "NO TO" in row_str.upper():
            continue

        print(f"  {filled}")
