import openpyxl
from openpyxl import Workbook

rows = []

# Helper
def add(question_id, variable_name, question_text, question_type, answer_option="", is_weight="No"):
    rows.append([question_id, variable_name, question_text, question_type, answer_option, is_weight])

# --- Hcountry (single) ---
q = "Country"
for val, label in [("1","KSA"),("2","UAE"),("3","Kuwait"),("4","Oman")]:
    add("Hcountry", "Hcountry", q, "single", f"{val}={label}")

# --- S1 (single) ---
q = "Please indicate if you or any member of your household are employed by ANY of the following."
for val, label in [("1","Manufacturing of goods and products"),("2","Banking / Finance"),("3","E-commerce"),
                   ("4","Real-estate"),("5","Owner of travel/tourism company/Travel agent"),
                   ("6","Market research company"),("7","Newspaper / TV company"),("8","Advertising agency"),("98","None of these")]:
    add("S1", "S1", q, "single", f"{val}={label}")

# --- S6 (single) ---
q = "What is your highest grade or level of education?"
for val, label in [("1","No formal schooling"),("2","Did not complete high school"),("3","SSC / HSC"),
                   ("4","Some college but did not graduate"),("5","Graduate"),("6","Postgraduate"),("98","Other")]:
    add("S6", "S6", q, "single", f"{val}={label}")

# --- S10c1 (single) ---
q = "Domestic / Within your country - In the past 12 months, how many times have you travelled for leisure holiday purposes?"
for val, label in [("1","Only once"),("2","Twice"),("3","3 times"),("4","4 to 6 times"),("5","More than 6 times"),("99","Didn't travel in last 12 months")]:
    add("S10c1", "S10c1", q, "single", f"{val}={label}")

# --- S10c2 (single) ---
q = "International - In the past 12 months, how many times have you travelled for leisure holiday purposes?"
for val, label in [("1","Only once"),("2","Twice"),("3","3 times"),("4","4 to 6 times"),("5","More than 6 times"),("99","Didn't travel in last 12 months")]:
    add("S10c2", "S10c2", q, "single", f"{val}={label}")

# --- S11 (single) ---
q = "When was the last time you travelled for leisure holiday purposes?"
for val, label in [("1","Last month"),("2","In the last 3 months"),("3","3 to 6 months ago"),
                   ("4","6 to 8 months ago"),("5","8 to 10 months ago"),("6","10 to 12 months ago"),("7","More than 12 months ago")]:
    add("S11", "S11", q, "single", f"{val}={label}")

# --- S12 (single) ---
q = "How likely are you to travel for leisure holiday purposes in the next 12 months?"
for val, label in [("1","Very Likely"),("2","Somewhat likely"),("3","Not sure right now"),("4","Not likely to travel")]:
    add("S12", "S12", q, "single", f"{val}={label}")

# --- S13a multi (regions visited past 3 years) ---
q = "Which of the following regions / countries have you visited in the past 3 years for leisure holiday purposes?"
s13a_vars = [
    ("S13a_Newr1","Western Europe"),("S13a_Newr2","Central Europe"),("S13a_Newr3","Eastern Europe"),
    ("S13a_Newr4","Asia"),("S13a_Newr5","Australia"),("S13a_Newr6","Africa"),("S13a_Newr7","Middle East"),
    ("S13a_Newr8","North America"),("S13a_Newr9","Central America"),("S13a_Newr10","South America"),
    ("S13a_Newr99","None of these")
]
for var, label in s13a_vars:
    add("S13a", var, q, "multi", label)

# --- S13r multi (regions never considered) ---
q = "Which of the following regions / countries would you never consider for leisure holiday purposes?"
s13r_vars = [
    ("S13r1","Western Europe"),("S13r2","Central Europe"),("S13r3","Eastern Europe"),
    ("S13r4","Asia"),("S13r5","Australia"),("S13r6","Africa"),("S13r7","Saudi Arabia"),
    ("S13r8","UAE"),("S13r9","North America"),("S13r10","Central America"),("S13r11","South America"),
    ("S13r12","I would consider all"),("S13r13","I would consider none")
]
for var, label in s13r_vars:
    add("S13r", var, q, "multi", label)

# --- Q1 multi (destinations heard of) ---
q = "Which of the following leisure holiday destinations within Saudi Arabia have you heard of? (Select all that apply)"
q1_vars = [
    ("Q1r1","AlUla"),("Q1r2","Neom City"),("Q1r3","Red Sea Project"),("Q1r4","Qiddiya"),
    ("Q1r5","Amaala"),("Q1r6","Jeddah"),("Q1r7","Medinna"),("Q1r8","Umluj"),("Q1r9","Abha"),
    ("Q1r10","Taif"),("Q1r11","Riyadh"),("Q1r12","Al Khobar"),("Q1r13","Dhahran"),("Q1r14","Dammam"),
    ("Q1r15","Yanbu"),("Q1r16","Tabuk"),("Q1r17","Khaybar"),("Q1r18","Tayma"),("Q1r19","NONE")
]
for var, label in q1_vars:
    add("Q1", var, q, "multi", label)

# --- Q2a multi (destinations ever travelled) ---
q = "Which of these leisure holiday destinations have you ever travelled? (Select all that apply)"
q2a_vars = [
    ("Q2ar1","AlUla"),("Q2ar2","Neom City"),("Q2ar3","Red Sea Project"),("Q2ar4","Qiddiya"),
    ("Q2ar5","Amaala"),("Q2ar6","Jeddah"),("Q2ar7","Medinna"),("Q2ar8","Umluj"),("Q2ar9","Abha"),
    ("Q2ar10","Taif"),("Q2ar11","Riyadh"),("Q2ar12","Al Khobar"),("Q2ar13","Dhahran"),("Q2ar14","Dammam"),
    ("Q2ar15","Yanbu"),("Q2ar16","Tabuk"),("Q2ar17","Khaybar"),("Q2ar18","Tayma"),("Q2ar19","NONE")
]
for var, label in q2a_vars:
    add("Q2a", var, q, "multi", label)

# --- Q2b multi (destinations likely to consider) ---
q = "Which of these destinations are you likely to consider for your leisure holiday in the next 12 months?"
q2b_vars = [
    ("Q2br1","AlUla"),("Q2br2","Neom City"),("Q2br3","Red Sea Project"),("Q2br4","Qiddiya"),
    ("Q2br5","Amaala"),("Q2br6","Jeddah"),("Q2br7","Medinna"),("Q2br8","Umluj"),("Q2br9","Abha"),
    ("Q2br10","Taif"),("Q2br11","Riyadh"),("Q2br12","Al Khobar"),("Q2br13","Dhahran"),("Q2br14","Dammam"),
    ("Q2br15","Yanbu"),("Q2br16","Tabuk"),("Q2br17","Khaybar"),("Q2br18","Tayma")
]
for var, label in q2b_vars:
    add("Q2b", var, q, "multi", label)

# --- Q10 scale ---
q = "How interesting or boring is this advertisement?"
for val, label in [("1","Very boring"),("2","Somewhat boring"),("3","Neither interesting nor boring"),
                   ("4","Somewhat interesting"),("5","Very interesting")]:
    add("Q10", "Q10_Lr1", q, "scale", f"{val}={label}")

# --- Q11 scale ---
q = "How easy or difficult is the advertisement to understand?"
for val, label in [("1","Very difficult"),("2","Somewhat difficult"),("3","Neither easy nor difficult"),
                   ("4","Somewhat easy"),("5","Very easy")]:
    add("Q11", "Q11_Lr1", q, "scale", f"{val}={label}")

# --- Q12 scale ---
q = "How much did you like the music in the ad?"
for val, label in [("1","Not liked it at all"),("2","Dislike it somewhat"),("3","Neither like it nor dislike it"),
                   ("4","Like it somewhat"),("5","Like it very much")]:
    add("Q12", "Q12_Lr1", q, "scale", f"{val}={label}")

# Write Excel
wb = Workbook()
ws = wb.active
ws.title = "Datamap"
ws.append(["question_id", "variable_name", "question_text", "question_type", "answer_option", "is_weight"])
for row in rows:
    ws.append(row)

# Style header
from openpyxl.styles import Font, PatternFill
for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor="4472C4")
    cell.font = Font(bold=True, color="FFFFFF")

# Auto column widths
for col in ws.columns:
    max_len = max(len(str(cell.value or "")) for cell in col)
    ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

out_path = r"C:\Users\KarthikShanmugam\ClaudePOC\DataLens\Module 1\data\spss\DataLens_Datamap.xlsx"
wb.save(out_path)
print(f"Datamap saved: {out_path}")
print(f"Total rows: {len(rows)}")
